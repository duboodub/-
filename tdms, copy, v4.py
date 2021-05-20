# -*- coding: utf-8 -*-
"""
Created on Tue Apr 13 16:27:28 2021

@author: db
"""


from openpyxl import Workbook
import os
import sys
from nptdms import TdmsFile
from nptdms import TdmsFile, TdmsWriter, RootObject, GroupObject, ChannelObject

def folder_path_tranform(f_path =''):
    folder_path = f_path.replace('\\', '/')
    return folder_path


def copy_from_tdms_to_xlsx(folder_path, file_name):
    wb = Workbook()
    
    with TdmsFile.open(os.path.join(folder_path,file_name), True) as tdms_file:
        for group in tdms_file.groups():
            g = group
            ws = wb.create_sheet(group.name)
            col=0
            for channel in group.channels():
                col = col+1
                ws.cell(1,col, channel.name)
                l = channel._length
                for i in range(0,l):
                    ws.cell(i+2,col, channel[i])
        file_name_ex = os.path.splitext(file_name)[0]
        wb.save(file_path_of_xlsx(folder_path, file_name_ex))
        ws.iter_cols(min_col=0, max_col=1, min_row=1,)
        

def file_path_of_tdms(folder_path, file_name):
     file_path_of_tdms = folder_path + '/' + file_name + '.tdms'
     return file_path_of_tdms
 
    
def file_path_of_xlsx(folder_path, file_name):
     file_path_of_xlsx = folder_path + '/' + file_name + '.xlsx'
     return file_path_of_xlsx



def list_files(folder_path,postsome='.tdms'):
    '''
    list like 'aaa.tdms, bbb.tdms'.

    Parameters
    ----------
    folder_path : TYPE
        DESCRIPTION.
    postsome : TYPE, optional
        DESCRIPTION. The default is '.tdms'.

    Returns
    -------
    file_list : TYPE
        DESCRIPTION.

    '''
    file_list = []
    
    for file in os.listdir(folder_path):
        if file.endswith(postsome):
            file_list.append(file)
            #file_list.append(os.path.join(folder_path,file))
            
    return file_list

 

if __name__ == '__main__':
    
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210418f1'
    # r'' 正则表达式，raw string， 少用
    folder_path = folder_path.replace('\\', '/')
    
    file_list = list_files(folder_path,)
    for file_name in file_list:
        print('%s copying.'%file_name)
        copy_from_tdms_to_xlsx(folder_path, file_name)
        print('%s is copied.'%file_name)
    print('all copied.')
    
    