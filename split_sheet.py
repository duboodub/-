# -*- coding: utf-8 -*-
"""
Created on Mon May 10 17:43:40 2021

@author: db
"""

import matplotlib.pyplot as plt 
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import os


def write_df_to_excel_as_additonal_sheet(df, excel_path:str, sheet_name:str):
    '''
    

    Parameters
    ----------
    df : TYPE
        DESCRIPTION.
    excel_path : TYPE
        DESCRIPTION.
    sheet_name : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    
    write df to an existing excel as a new sheet, or replace the sheet with a same sheet_name
    '''
    wb = load_workbook(excel_path)
    s = wb.sheetnames
    if sheet_name not in wb.sheetnames:
        with pd.ExcelWriter(excel_path,engine= 'openpyxl') as writer:
            writer.book = wb
            # writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
            
            # wb = load_workbook('sample_file.xlsx')
            # with pd.ExcelWriter('sample_file.xlsx',engine= 'openpyxl') as writer:
            # 如两个workbook一样则会新创一个sheet，如不一样这直接修改
                
            df.to_excel(writer, sheet_name, index= False)
            #如sheet name冲突，则会新创一个sheet
            
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
            writer.save()
    
    else:
        pass
    
    
def create_ploting_sheet_1(excel_path:str, sheet_name:str, temp_list:list,):
    '''
    RHT, 2450 only 
    temp_list: list of temperature, RHT_2450
    sheet_name_source = 'Scan H,2450'
    sheet_name_new = 'T=%dK %s' %(t, sheet_name)
            
    '''
    # temp_list = [295, 285, 275]
    # temp_list = np.linspace(5,295,30,dtype=int)
    
    df =   pd.read_excel(excel_path, sheet_name, engine= 'openpyxl')
    
    for t in temp_list:  
        print('%dK, sheet creating\t'%t)
        df1 = df[( (t-1) < df['sample temperature/K'] ) & ( df['sample temperature/K'] < (t+1) )]
               # https://pandas.pydata.org/docs/user_guide/indexing.html
               # Another common operation is the use of boolean vectors to filter the data. The operators are: | for or, & for and, and ~ for not.
        
        write_df_to_excel_as_additonal_sheet(df1, excel_path, 'T=%dK %s' %(t, sheet_name), )
        print('%dK, done'%t)
        
    return



if __name__ == '__main__':
    # modify
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210404f1'
    # r'' 正则表达式，raw string， 少用
    
    excel_name = 'DB20210404m6,.xlsx'
    excel_path = os.path.join(folder_path, excel_name)
    
    sheet_name_source = 'Scan H,2450'
    # sheet_name_new = 'T=%dK %s' %(t, sheet_name), e.g. 'T=300K sheet_name_source'
    
    temp_list = np.linspace(5,295,10,dtype=int)     # temp_list = np.linspace(5,295,10,dtype=int) = [5,15,...,285,295]
    
    col_x = 'Lakeshore643 Output/A'
    col_y = 'Resistance(Ω) (2450)'
    
    
    folder_path = folder_path.replace('\\', '/')
    excel_path = os.path.join(folder_path, excel_name)
    
    create_ploting_sheet_1(excel_path,sheet_name_source,temp_list,)

