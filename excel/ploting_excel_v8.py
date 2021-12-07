# -*- coding: utf-8 -*-
"""
Created on Sat Apr 17 03:05:53 2021

@author: db
"""




#pf.plot
  
import matplotlib.pyplot as plt 
import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np
import os

class measure_file():
    
    def __init__(self, folder_path, file_name):
        self.folder_path = folder_path
        self.file_name = file_name
        self.file_nm = os.path.splitext(file_name)[0]
        self.file_post = os.path.splitext(file_name)[1]
        
        
class measure_excel(measure_file):
    
    def __init__(self, folder_path, file_name):
        measure_file.__init__(self, folder_path, file_name)
        self.measure_id = self.find_measure_id()
        super()
        
        if self.file_post != 'xlsx':
            print('It\'s not a excel, you can\'t assign this file as measure_excel!')
            return 1
         
        
        
    def find_measure_id(self):
        s0 = self.file_name.find('DB')
        s1 = s0 + 8
        measure_id = self.file_name[s0:s1]
        return measure_id
    
class data_sheet(measure_excel):
    def __init__(self, folder_path, file_name, sheet_name):
        measure_excel.__init__(self, folder_path, file_name)
        self.sheet_name = sheet_name
        self.data = self.to_df()
        
    def to_df(self):
        pd.read_excel(self.excel_path, sheet_name=self.sheet_name, engine='openpyxl',)
        

def write_df_to_excel_as_additonal_sheet(df, excel_path:str, sheet_name:str):
    '''
    

    Parameters
    ----------
    df : TYPE
        DESCRIPTION.
    excel_path : TYPE
        DESCRIPTION.
    sheet_name : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    
    write df to an existing excel as a new sheet, or replace the sheet with a same sheet_name
    '''
    
    if os.path.exists(excel_path):
        pass
    else:
        print('excel_path doesn\'t exist! So a new file created.\n')
        wb = Workbook()
        wb.save(excel_path)
        
    wb = load_workbook(excel_path)
    
    s = wb.sheetnames
    if sheet_name not in wb.sheetnames:
        with pd.ExcelWriter(excel_path,engine= 'openpyxl') as writer:
            writer.book = wb
            # writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
            
            # wb = load_workbook('sample_file.xlsx')
            # with pd.ExcelWriter('sample_file.xlsx',engine= 'openpyxl') as writer:
            # 如两个workbook一样则会新创一个sheet，如不一样这直接修改
                
            df.to_excel(writer, sheet_name, index= False)
            #如sheet name冲突，则会新创一个sheet
            
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
            writer.save()
            
    
    else:
        print('{} exists! So passed.\n'.format(sheet_name))
    
    
#PXI
def create_ploting_sheet_0(excel_path, sheet_name_source):
    '''
        PXIe
        copied excel -> calculate R of each channel  -> add ploting sheet
        __excel_path__ = 2
        
    '''
    sheet_name_new = 'R %s'%sheet_name_source
    
    R_s2_ch2   = 10000
    R_s3_ch2  = 10000
    # I = dict(I_ch0 = I_ch3)

    
    df0 =   pd.read_excel(excel_path, sheet_name_source )

    
    # df1 = df0.iloc[:,:35]
    df1 = df0
        #df1改变df0依然跟着变，（指向同一储存内容）
    
    I_s2_ch2 = df1['CH2(slot2)'] /R_s2_ch2
    I_s3_ch2 = df1['CH2(slot3)'] /R_s3_ch2
    # I_s3_ch2 = df1.iloc[:,21] /R_s3_ch2
    
    #calculate R
    for i in range(16):
        if i not in [2,6,10,14]:
            df1.iloc[:,3+i] = df1.iloc[:,3+i] / I_s2_ch2
            df1.iloc[:,3+16+i] = df1.iloc[:,3+16+i] / I_s3_ch2
            
    write_df_to_excel_as_additonal_sheet(df1, excel_path, sheet_name_new) 

    return 0

    
#
def create_ploting_sheet_RHT(excel_path:str, sheet_name:str, temp_list:list,):
    '''
    RHT,  
    temp_list: list of temperature, RHT_2450
    sheet_name_source = 'Scan H,2450'
    sheet_name_new = 'T=%dK %s' %(t, sheet_name)
            
    '''
    # temp_list = [295, 285, 275]
    # temp_list = np.linspace(5,295,30,dtype=int)
    excel_path_tuple = os.path.splitext(excel_path)
    excel_path_new ='%s, %s%s'%(excel_path_tuple[0], sheet_name, excel_path_tuple[1])
    
    df =   pd.read_excel(excel_path, sheet_name, engine= 'openpyxl')
    
    write_df_to_excel_as_additonal_sheet(df, excel_path_new, sheet_name, )
    
    for t in temp_list:  
        print('%s %dK, sheet creating,\t'%(sheet_name,t))
        # df1 = df[( (t-1) < df['sample temperature/K'] ) & ( df['sample temperature/K'] < (t+1) )]
               # https://pandas.pydata.org/docs/user_guide/indexing.html
               # Another common operation is the use of boolean vectors to filter the data. The operators are: | for or, & for and, and ~ for not.
        
        write_df_to_excel_as_additonal_sheet(df[( (t-1) < df['sample temperature/K'] ) & ( df['sample temperature/K'] < (t+1) )],
                                             excel_path_new, 
                                             'T=%dK %s' %(t, sheet_name), )
        print('%s %dK, sheet created!\t'%(sheet_name,t))
        
    return


# 
def plot_template_PXI_T(excel_path,sheet_name_source, col_x, col_y ):
    '''
        single curve
        col_x: column index in the sheet (df)
        
    

    Parameters
    ----------
    excel_path : TYPE
        DESCRIPTION.
    sheet_name_source : TYPE
        DESCRIPTION.
    col_x : TYPE
        DESCRIPTION.
    col_y : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    sheet_name = sheet_name_source

    #    col_x = 0
    # col_y = column_of_channel_0(2, 1)
    
    # pt= df1.plot(x=df1.columns[0], y = df1.columns[column_of_channel_0(0,2):column_of_channel_0(15,2)], subplots=True, figsize= None )
    # create_ploting_sheet_0(excel_path,sheet_name_source,sheet_name_new)
    df =   pd.read_excel(excel_path, sheet_name, engine= 'openpyxl')
    pl =  df.plot(x=df.columns[col_x], y= df.columns[col_y], figsize=(25,20))
    ch = df.columns[col_y] 
    # ch1 = df.columns[6:10]
    font1 = {'size': 20}
    plt.title(excel_path,fontdict=font1)
    plt.ylabel(col_y,fontdict=font1)
    plt.xlabel(col_x,fontdict=font1)
    plt.grid(which='major')  # 生成网格
    plt.legend(prop=font1)  # 显示legend
    
    # qtn, save fig
    fig = pl.get_figure()
    fig.savefig(ch +'.png')
   

#
def sheet_names_from_Ts(temp_list):     
    sheet_name_list = []
    for T in temp_list:
        # sheet_name = 'T=%dK %s' %(T,sheet_name_original)
        sheet_name = 'T=%dK 2450 only data_1'%T   # 2450 only
        # sheet_name = 'T=%dK_(Scan H,2450)'%T    # DB20210404m5,.xlsx
        sheet_name_list.append(sheet_name)
        
    return sheet_name_list
        


def plot_template_RHT_2450(excel_path, sheet_name_list, x_name, y_name):
    '''
    
    all in a figure
    
    sheet_name_original: sheet_name = 'T=%dK %s' %(T,sheet_name_original)
    
    for 2450 only:
    x_name = 'Lakeshore643 Output/A'
    y_name = 'Resistance(Ω) (2450)'
    
    for switch:
        x_name = 'Lakeshore643 Output/A'
        y_name ='Resistance(Ω) '
    
    '''
    
    # excel_path = 'DB20210404m5,.xlsx'
    excel_name = excel_path
    # sheet_name = 'T=205K_(Scan H,2450)'
    # temp_list = np.linspace(295,5,30,dtype=int)
    
    

    figurreNumber = 0
    
    for sheet_name in sheet_name_list:
        df =   pd.read_excel(excel_path, sheet_name, engine= 'openpyxl')
        
        

        fn = plt.figure(num=figurreNumber, figsize=(20,20),)  #新建一个pic,  https://blog.csdn.net/haikuotiankong7/article/details/90551841
        
        pt = plt.plot(df[x_name], df[y_name], label=sheet_name,)
        # plt.plot(otherFileData[xLabelVar], otherFileData[checkedVlaueName], '--b', label="other")
        
        font1 = {'size': 20}
        plt.title(os.path.split(excel_path)[-1],fontdict=font1)
        plt.ylabel(y_name,fontdict=font1)
        plt.xlabel(x_name,fontdict=font1)
        plt.grid(which='major')  # 生成网格
        plt.legend( loc='center left', bbox_to_anchor=(1,0.5), prop=font1)  # 显示legend
        plt.savefig('%s/RHT2_%s.png' %(os.path.split(excel_path)[0],os.path.split(excel_path)[-1]), dpi=300, bbox_inches='tight')
        # fn.clear1
        #plt.show() # 逐个显示一个图形，点击关闭后继续显示另一个
    # plt.show()  # 将所有的图形一下显示出来
        # plt.legend(loc='best', bbox_to_anchor=(1, 0.5), prop=font1)  # 显示legend

        # qtn, save fig
        # fig = pl.get_figure()
        # fig.savefig(sheet_name +'.png')
    pass


# temp
def RHT_2450():
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210404f1'
    # r'' 正则表达式，raw string， 少用
    folder_path = folder_path.replace('\\', '/')
    
    excel_name = 'DB20210404m6,.xlsx'
    excel_path = os.path.join(folder_path, excel_name)
    excel_nm = os.path.splitext(excel_name)[0]
    # sheet_name_source = 'T=300K 2450 only data_1'
    

    # sheet_name_source = 'Scan H,2450'
    # sheet_name_new = 'T=205K Scan H,2450'
    
    temp_list = np.linspace(295,5,10,dtype=int)
    # create_ploting_sheet_1(excel_path, sheet_name_source, temp_list,)
    
    col_x = 'Lakeshore643 Output/A'
    col_y = 'Resistance(Ω) (2450)'
    
    create_ploting_sheet_RHT(excel_path,sheet_name_source,temp_list,)
    plot_template_RHT_2450(excel_path, 'sheet_name_list', col_x, col_y)
    
    
# temp
def PXI():
    excel_path = 'DB20210404m3.xlsx'
    sheet_name_source = 'new PXI data+'
    sheet_name_R = 'R %s'%sheet_name_source
    # create_ploting_sheet_1(excel_path, sheet_name_source, temp_list,)


    col_x = 0
    cy_xy = [4,12]
    cy_xx = [1,3,5,7,8,9,11]
    
    slot_id = 2
    col_xx = []
    for i in cy_xx:
        col_i = column_of_channel_0(slot_id, i)
        col_xx.append(col_i)
        
    col_xy = []
    for i in cy_xy:
        col_i = column_of_channel_0(slot_id, i)
        col_xy.append(col_i)
        
        
        
    col_y = column_of_channel_0(2,4)

    #    col_x = 0
    # col_y = column_of_channel_0(2, 1)
    
    # pt= df1.plot(x=df1.columns[0], y = df1.columns[column_of_channel_0(0,2):column_of_channel_0(15,2)], subplots=True, figsize= None )
    # create_ploting_sheet_0(excel_path,sheet_name_source,sheet_name_new)
    
    df =   pd.read_excel(excel_path, sheet_name_R, engine= 'openpyxl')
    
    
    
    
    pl =  df.plot(x=df.columns[col_x], y= df.columns[col_xx], figsize=(25,20))
    ch = df.columns[col_y]
    
    # ch1 = df.columns[6:10]
    font1 = {'size': 20}
    plt.title(excel_path,fontdict=font1)
    plt.ylabel('R/Ohm',fontdict=font1)
    plt.xlabel(df.columns[col_x],fontdict=font1)
    plt.grid(which='major')  # 生成网格
    plt.legend(prop=font1)  # 显示legend
    
    # qtn, save fig
    fig = pl.get_figure()
    fig.savefig( '%s,2Rxx.png'%excel_path,)
    fig.clear('all')
    
    
    
    
    pl =  df.plot(x=df.columns[col_x], y= df.columns[col_xy], figsize=(25,20))
    ch = df.columns[col_y]
    
    # ch1 = df.columns[6:10]
    font1 = {'size': 20}
    plt.title(excel_path,fontdict=font1)
    plt.ylabel('R/Ohm',fontdict=font1)
    plt.xlabel(df.columns[col_x],fontdict=font1)
    plt.grid(which='major')  # 生成网格
    plt.legend(prop=font1)  # 显示legend
    
    # qtn, save fig
    fig = pl.get_figure()
    fig.savefig( '%s,2Rxy.png'%excel_path, )
    fig.clear('all')
    
    
    
    col_y = column_of_channel_0(2,4)
    
    pl =  df.plot(x=df.columns[col_x], y= df.columns[col_y], figsize=(25,20))
    ch = df.columns[col_y]
    
    # ch1 = df.columns[6:10]
    font1 = {'size': 20}
    plt.title(excel_path,fontdict=font1)
    plt.ylabel('R/Ohm',fontdict=font1)
    plt.xlabel(df.columns[col_x],fontdict=font1)
    plt.grid(which='major')  # 生成网格
    plt.legend(prop=font1)  # 显示legend
    
    # qtn, save fig
    fig = pl.get_figure()
    fig.savefig( '%s,21Rxy.png'%excel_path,dpi=300)
    fig.clear('all')
    
    
     
def correct():
    '''
    PXI, correct the V which overflows
    

    Returns
    -------
    None.

    '''
    excel_path = 'DB20210404m3.xlsx'
    sheet_name_source = 'PXI data+'

    
    df =   pd.read_excel(excel_path, sheet_name_source, engine= 'openpyxl')
    
    bl = df['CH2(slot2)'] > 4.9
    ids = df.index[bl]
    for idx in ids:
        d1= df.loc[idx,'CH2(slot2)'] = 9 - df.loc[idx,'CH0(slot2)']
        
 
    bl = df['CH0(slot2)'] > 4.9
    ids = df.index[bl]
    for idx in ids:
        d1= df.loc[idx,'CH0(slot2)'] = 9 - df.loc[idx,'CH2(slot2)']
        
    write_df_to_excel_as_additonal_sheet(df,excel_path,'new %s'%sheet_name_source)
        
        

def column_of_channel_0(slot_id,channel_id):
    '''for pxi sheet'''
    
    if channel_id in range(16) and slot_id in [2, 3]:
        col_id = 3+ channel_id + (slot_id - 2)*16 
        return col_id

    else:
        return 'channel_id or slot_id Fault'


#
def creat_organized_sheet_2450(excel_path, sheet_name_source):
    '''
    combine several sheets (magnet output) to one for ploting

    Parameters
    ----------
    excel_path : TYPE
        DESCRIPTION.
    sheet_name_source : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210410f1\xlsx'
    # r'' 正则表达式，raw string， 少用
    folder_path = folder_path.replace('\\', '/')
    
    excel_name = 'DB20210410m12,RH,300k.xlsx'
    excel_path = os.path.join(folder_path, excel_name)
    
    sheet_name_source = '2450 only data'
    sheet_name_new = '%s_1' %sheet_name_source
    
    df = pd.read_excel(excel_path,sheet_name= sheet_name_source,engine='openpyxl')
    df1 = pd.read_excel(excel_path,sheet_name= 'magnet data',engine='openpyxl')
    df.insert(0,df1.columns[0],df1.iloc[:,0])
    
    write_df_to_excel_as_additonal_sheet(df, excel_path, sheet_name_new)
    
    # temp_list = np.linspace(10,300,30,dtype=int)
    # create_ploting_sheet_1(excel_path,sheet_name_new,temp_list)
    # plot_template_RHT_2450(excel_path,temp_list, sheet_name_new)
    
    pass


#
def creat_measure_sheets(excel_path:str):
    '''
    

    Parameters
    ----------
    excel_path : str
        DESCRIPTION.

    Returns
    -------
    sheet_name_new_list : TYPE
        DESCRIPTION.

    '''
    sheet_name_source_list = switch_measure_sheet_list(excel_path)
    sheet_name_new_list = creat_measure_sheet(excel_path, sheet_name_source_list)
    
    return sheet_name_new_list
    
'''    
def switch_measure_sheet_list(slot_channel_dict={1:[1,2,3], 2:[1,2,3]})->list:
    sheet_name_source_list = []
    for slot, channel_list in slot_channel_dict:
        for ch in channel_list:
            sheet_name_source = 'Switch-Slot%d CH%d'%(slot,ch)
            sheet_name_source_list.append(sheet_name_source)
    
    return sheet_name_source_list
'''
#
def switch_measure_sheet_list(excel_path:str):
    '''
    Parameters
    ----------
    excel_path : str
        DESCRIPTION.

    Returns
    -------
    sheet_name_source_list : TYPE
        DESCRIPTION.

    '''
    wb = load_workbook(excel_path,read_only=True,)
    sheet_name_source_list = []
    for sheetname in wb.sheetnames:
        if sheetname[:6] == 'Switch':
            sheet_name_source_list.append(sheetname)
    wb.close
    
    return sheet_name_source_list
    
#             
def creat_measure_sheet(excel_path:str, sheet_name_source_list:list, )->list:
    '''
    sheet_name_to_copy= 'magnet data'
    
    Parameters
    ----------
    excel_path : str
        DESCRIPTION.
    sheet_name_source_list : list
        DESCRIPTION.
     : TYPE
        DESCRIPTION.
        
    Returns
    -------
    list
        DESCRIPTION.
        
    '''
    # sheet_name_source = '2450 only data'
    sheet_name_to_copy= 'magnet data'
    # sheet_name_new = '%s_1' %sheet_name_source
    df1 = pd.read_excel(excel_path,sheet_name= sheet_name_to_copy, engine='openpyxl')
    
    # add magnet data column
    sheet_name_new_list = []
    for sheet_name_source in sheet_name_source_list:        
        sheet_name_new = '%s.ms' %sheet_name_source         # measure sheet
        sheet_name_new_list.append(sheet_name_new)
        df = pd.read_excel(excel_path,sheet_name= sheet_name_source,engine='openpyxl')
        # df.insert(0,df1.columns[0],df1.iloc[:,0])
        df.insert(0,df1.columns[0],df1.iloc[:,0])        
        write_df_to_excel_as_additonal_sheet(df, excel_path, sheet_name_new)
        print('1 measure sheet created\r')
    print('All measure sheets created\n')
    return sheet_name_new_list
    # creat plot sheet
    
    # temp_list = np.linspace(190,210,3,dtype=int)
    # for ch in channel_list:
    #     sheet_name_source = 'Switch-Slot1 CH%d'%ch

    #     sheet_name_new = '%s_1' %sheet_name_source
        
    #     create_ploting_sheet_1(excel_path,sheet_name_new,temp_list)
    #     print('ch%d done'%ch)
        
    # # plot
    # x_name = 'Lakeshore643 Output/A'
    # y_name ='Resistance(Ω) '
    # plot_template_RHT_2450(excel_path,temp_list, sheet_name_new, x_name, y_name)

    # pass


#
def df_section(df,H_max, H_step, N_of_section):
    '''
    RxyH-T, obtain RH

    Parameters
    ----------
    df : TYPE
        DESCRIPTION.
    H_max : TYPE
        DESCRIPTION.
    H_step : TYPE
        DESCRIPTION.
    N_of_section : TYPE
        DESCRIPTION.

    Returns
    -------
    df_N : TYPE
        DESCRIPTION.

    '''
    n = int(H_max/ H_step) + 1  # number in the section
    n1= (N_of_section-1) * n
    
    df_N = df[n1:(n1+n)] 
    
    return df_N


#
def df_combine(df,H_max:float, H_step:float, list_of_N:list, list_convert:list):
    '''
    将多段curve拼接成一个完整的 RH loop

    Parameters
    ----------
    df : TYPE
        DESCRIPTION.
    H_max : float
        DESCRIPTION.
    H_step : float
        DESCRIPTION.
    list_of_N : list
        DESCRIPTION.
    list_convert : list
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    '''
    if len(list_of_N) != len(list_convert):
        return 'len(list_of_N) != len(list_convert)'
    
    elif len(list_of_N) == len(list_convert):
        i = 0
        for N_of_section in list_of_N:
            if i == 0:
                df1 = df_section(df,H_max, H_step, N_of_section)
                
                if list_convert[i]:
                    df1 = df1[::-1]
                    
                
            else:
                df2 = df_section(df,H_max, H_step, N_of_section)
                if list_convert[i]:
                    df2 = df2[::-1]
                    
                df1 = df1.append(df2)
                
            i+=1
            
        return df1
        


#
def RHT_re_curve():
    '''
    去除开头不属于loop部分，构造一个loop循环

    Returns
    -------
    None.

    '''
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210404f1'
    # r'' 正则表达式，raw string， 少用
    folder_path = folder_path.replace('\\', '/')
    
    excel_name = 'DB20210404m6,.xlsx'
    excel_path = os.path.join(folder_path, excel_name)
    excel_nm = os.path.splitext(excel_name)[0]
    # sheet_name_source = 'T=300K 2450 only data_1'
    
    temp_list = np.arange(5,296,10)
    sheet_name_list = sheet_names_from_Ts(temp_list,)
    channel_list = [1,2,3,4]
    
    d = 1.32e-6
    x_name = 'Lakeshore643 Output/A'
    y_name = 'Resistance(Ω) (2450)'

    for sheet_name_source in sheet_name_list:
        
    
    
        df = pd.read_excel(excel_path,sheet_name= sheet_name_source,engine='openpyxl')
        
        # df_R = df_combine(df,40,0.5,[5,2,3,4],[0,0,0,0])
        # df_f = df_combine(df,40,0.5,[2,2,4,4],[1,0,1,0])
        
        # df_R = df_combine(df,50,0.25,[1,2,3,4],[0,0,0,0])
        # df_f = df_combine(df,50,0.25,[2,2,4,4],[1,0,1,0])

        df_R = df_combine(df,65,0.5,[1,2,3,4],[0,0,0,0])
        df_f = df_combine(df,65,0.5,[2,2,4,4],[1,0,1,0])        
        
        # df_f = df_combine(df,50,0.25,[1,1,3,3],[0,1,0,1])
        
        df_f.plot(x= x_name, y=y_name)
        I_H = df_f[x_name].to_numpy()
        R_f = df_f[y_name].to_numpy()
        R_R = df_R[y_name].to_numpy()
        R_hysteresis = R_R - R_f
        R_fc = R_f[::-1]
        R_h = 1/2 * (R_f - R_fc)
        R_g = 1/2 * (R_f + R_fc)
        f = np.polyfit(I_H, R_h, 9)
        linfit = np.polyval(f, I_H)
        c=len(df_R.index)
        K = np.full(len(df_R.index), f[0])
        RH = K * d
        
        df_R.insert(len(df_R.columns), 'R_f', R_f)                       # R_f: 没有loop的曲线
        df_R.insert(len(df_R.columns), 'R_g', R_g)                       # R_g: R_f分解的偶函数
        df_R.insert(len(df_R.columns), 'R_h', R_h)                       # R_h: R_f分解的奇函数（hall）
        df_R.insert(len(df_R.columns), 'R_hysteresis', R_hysteresis)      # R_hysteresis: R_f的hysteresis成分
        df_R.insert(len(df_R.columns), 'linfit', linfit)                  # linfit: 对R_h的多项式拟合
        df_R.insert(len(df_R.columns), 'K', K)                           # 拟合的一次项
        df_R.insert(len(df_R.columns), 'RH', RH)                        # hall系数
        
        write_df_to_excel_as_additonal_sheet(df_R, excel_path,'RH %s'%sheet_name_source)
        print('one sheet created.')

        
    for sheet_name_source in sheet_name_list:
        sheet_name_new = 'RH %s'%sheet_name_source
        df_R = pd.read_excel(excel_path,sheet_name= sheet_name_new,engine='openpyxl')
        
        
        
        # plot
        fn = plt.figure(num=1, figsize=(20,20),)  #新建一个pic,  https://blog.csdn.net/haikuotiankong7/article/details/90551841
        # fn = plt.figure(num=1, figsize=(20,20),clear=True)  #新建一个pic,  https://blog.csdn.net/haikuotiankong7/article/details/90551841

        # pt = plt.plot(I_H, linfit, label='linfit',)
        # pt = plt.plot(I_H, R_g, label='R_g',)
        # pt = plt.plot(I_H, R_h, label='R_h',)
        # pt = plt.plot(I_H, linfit,)
        pt = plt.plot(df_R[x_name], df_R['R_h'],label= sheet_name_new)
        # pt = plt.plot(df[x_name], df[y_name],label=y_name)
        # plt.plot(otherFileData[xLabelVar], otherFileData[checkedVlaueName], '--b', label="other")
        
    font1 = {'size': 20}
    plt.title('R_h' + os.path.split(excel_path)[-1],fontdict=font1)
    plt.ylabel('R_h',fontdict=font1)
    plt.xlabel(x_name,fontdict=font1)
    plt.grid(which='major')  # 生成网格
    plt.legend( loc='center left', bbox_to_anchor=(1,0.5), prop={'size': 14})  # 显示legend
    plt.savefig('%s/R_h(245k-295k) %s %s.png' %(folder_path, sheet_name_new, os.path.splitext(os.path.split(excel_path)[-1])[0]), dpi=300, bbox_inches='tight')
    print('one plotted')

       
         

if __name__ == '__main__':
    
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210518f1\excel_copied_from_tdms'
    # r'' 正则表达式，raw string， 少用
    folder_path = folder_path.replace('\\', '/')
    
    excel_name = 'DB20210518m10， LSMO 5+65，.xlsx'
    excel_path = os.path.join(folder_path, excel_name)
    excel_nm = os.path.splitext(excel_name)[0]
    
    
    ms_name_list = creat_measure_sheets(excel_path)
    temp_list = np.arange(210,230, 10)
    #temp_list= [5,10,20,30,40]
    
    for ms_name in ms_name_list:
        create_ploting_sheet_RHT(excel_path,ms_name,temp_list)
        
        
        
    # for sheet_name_source in sheet_name_list:
    #     sheet_name_new = 'RH %s'%sheet_name_source
    #     df_R = pd.read_excel(excel_path,sheet_name= sheet_name_new,engine='openpyxl')
        
        
        
    #     # plot
    #     fn = plt.figure(num=1, figsize=(20,20),)  #新建一个pic,  https://blog.csdn.net/haikuotiankong7/article/details/90551841
    #     # fn = plt.figure(num=1, figsize=(20,20),clear=True)  #新建一个pic,  https://blog.csdn.net/haikuotiankong7/article/details/90551841

    #     # pt = plt.plot(I_H, linfit, label='linfit',)
    #     # pt = plt.plot(I_H, R_g, label='R_g',)
    #     # pt = plt.plot(I_H, R_h, label='R_h',)
    #     # pt = plt.plot(I_H, linfit,)
    #     pt = plt.plot(df_R[x_name], df_R['R_h'],label= sheet_name_new)
    #     # pt = plt.plot(df[x_name], df[y_name],label=y_name)
    #     # plt.plot(otherFileData[xLabelVar], otherFileData[checkedVlaueName], '--b', label="other")
        
    # font1 = {'size': 20}
    # plt.title('R_h' + os.path.split(excel_path)[-1],fontdict=font1)
    # plt.ylabel('R_h',fontdict=font1)
    # plt.xlabel(x_name,fontdict=font1)
    # plt.grid(which='major')  # 生成网格
    # plt.legend( loc='center left', bbox_to_anchor=(1,0.5), prop={'size': 14})  # 显示legend
    # plt.savefig('%s/R_h(245k-295k) %s %s.png' %(folder_path, sheet_name_new, os.path.splitext(os.path.split(excel_path)[-1])[0]), dpi=300, bbox_inches='tight')
    # print('one plotted')

    