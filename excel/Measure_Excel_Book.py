# -*- coding: utf-8 -*-
"""
Created on Sat Jun 12 16:16:05 2021

@author: db
"""
import matplotlib.pyplot as plt 
import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np
import os


from data_process_oop import Excel_Book, Excel_Sheet, File

class Measure_Excel_Book(Excel_Book):
    
    def __init__(self,folder_path, file_name, measure_id,):
        super().__init__(self, folder_path, file_name, )
        self.measure_id = measure_id        ## 
        
    ##    
    def load_measure_excel_book(self, folder_path, file_name, read_only:bool = False, measure_id='',):
        
        super().load_excel_book( self, folder_path=folder_path, file_name=file_name, read_only=read_only)       # ?self
        self.measure_id = measure_id
        
    
    # same def save(self, folder_path= '', file_name='', )
    
    
    
def process_RHT():
    '''
    input:  folder_path, 
            H_dir, 
            switch_channel_list, 
            para_cal_rho_cm_list, 
            mea_id_list, 
            temperatures_list,
            dT

    Returns
    -------
    None.

    '''
    
    # input folder_path
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20211125f1, LCO-LSMO0.67, RT, RHxyT\excel_copied_from_tdms\RHT\RHzT'
    folder_path_m = folder_path + r'\meb'           # measure books
    
    H_dir = 'z'

    '''
    create meb
        input folder_path
        creat measure excel books of the excel books in the folder_path. concat all the columns.
        so far measure_type='RHT,switch'  only.
        
    '''
    file_name_list= os.listdir(folder_path)
    
    path_dir = folder_path_m
    if not os.path.exists(path_dir):
        print('dir of meb doesn\'t exist, a new one created!\n')
        os.makedirs(path_dir)
        
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1] == '.xlsx':
            print('creating measure excel book of {}'.format(file_name))
            eb = Excel_Book.load_excel_book(folder_path,file_name)
            eb.write_to_measure_excel_book(d_folder_path=folder_path_m, d_file_name='', measure_type='RHT,switch')
            print('created measure excel book of {}\n'.format(file_name))
    
    
    
    '''
    insert H calculated from I(magnet).
        input folder_path_m, H_dir (I_to_H(H_dir)) H_dir = 'z' or 'xy'
        在folder_path目录下的excel文件中插入H列，由定标曲线拟合从I(magnet)转换得到。
        设置 folder_path_m, 即可插H。
    '''
    print('\ninserting H calculated from I(magnet).')
    
    # input H_dir
    # H_dir = 'z'
    
    file_name_list = os.listdir(folder_path_m)
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1]=='.xlsx':
            print('\nprocessing excel book: {}'.format(file_name))
            eb = Excel_Book.load_excel_book(folder_path_m,file_name)
            for sheet_name in eb.wb.sheetnames:
                es = Excel_Sheet(eb, sheet_name)
                if 'Lakeshore643 Output/A' in es.df_es.columns:
                    print('{}.H inserting.'.format(sheet_name))
                    es.I_to_H(H_dir)
                    es.save()
    
    
    '''
    calculate \rho & \rho_sheet of meb
    input para_cal_rho_cm_list [length, width, thickness],
    '''
    print('\ncalculating rho & rho_sheet of meb.')
    file_name_list = os.listdir(folder_path_m)
    
    switch_channel_list = [ 
                            [1,1],[1,2],[1,3],[1,4],[1,5],[1,6],[1,7],[1,8],[1,9],[1,10],
                            [2,1],[2,2],[2,3],[2,4],[2,5],[2,6],[2,7],[2,8],[2,9],[2,10],
                          ]
          
    # input para_cal_rho_cm_list: [
    #       para_cal_rho_cm_list = [
    #                                 (slot1 channel1): [length width thickness],
    #                                 ...
    #                                 (slot1 channel10): [length width thickness],
    #                                 (slot2 channel1): [length width thickness],
    #                                 ...
    #                                 (slot2 channel10): [length width thickness],
    #                             ]

    para_cal_rho_cm_list = [
                              [],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                                 
                              [],                                  

                              [3e-2, 1e-2, 2.6e-6], 
                              [],
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [], 
                              
    
                              [],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                                 
                              [],                                  

                              [3e-2, 1e-2, 2.6e-6], 
                              [],
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                               
                            ]
    
    para_list = list(zip(switch_channel_list, para_cal_rho_cm_list))  
    
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1]=='.xlsx':
            print(file_name +': rho + rho_sheet')
            eb= Excel_Book.load_excel_book(folder_path_m, file_name)
            for para in para_list:
                if not para[1] == []:
                    channel_name = 'Switch-Slot%d CH%d'%(para[0][0], para[0][1])
                    
                    for sheet_name in eb.wb.sheetnames:
                        if channel_name in sheet_name:
                            es = Excel_Sheet(eb, sheet_name)
                            print(es.sheet_name)
                            l=para[1]
                            es.calculate_rho(l)
                            es.calculate_rho(l[0:2])
                            es.save()
                            
                    
                    
    '''
    create peb
        input folder_path_m, file_name_list, temperatures_list (of the temperatures in the excel books of RHT),
        
        split sheets of different T of one channel (sheet_name), and save with file_name = <'t=p, ' + self.sheet_name +'.xlsx'>, 
            sheet_name = <T={}K,{},{}'.format(temperature, self.sheet_name, self.file.file_nm[5:20])> if not specified.
            
    '''
    #input mea_id_list
    mea_id_list =   [
                        'DB20211125f1m39, RHzT',
                        'DB20211129f1m2, LCO-LSMO0.67, RHzT',
                        'DB20211129f1m5, LCO-LSMO0.67, RHzT',


                        ]
                    #     'DB20210510m27, RHT, switch, 190k-5k.xlsx',
                    #     'DB20210510m28, RHT, switch, 170k-.xlsx',
                    #     'DB20210510m33, RHT, switch, 40k-.xlsx',
                    #     'DB20210510m35, RHT, switch, .xlsx',
                        
                    # ]
    file_names = os.listdir(folder_path_m)
    file_name_list = []
    for mea_id in mea_id_list:
        for file_name in file_names:
            if mea_id in file_name:
                file_name_list.append(file_name)
    
    
    # input temperatures_list
    tl =  list(np.arange(120,301, 10))
    tl = tl.append(320)
    temperatures_list = [   [60,80,90, 100],
                            [300],
                            [310, 320],
                            ]
                        #   np.arange(320,199,-10),
                        #   [190,180],
                        #   np.arange(170,39,-10),
                        #   np.arange(40,9,-10),
                        #   [320],                            
                        # ]
    
    folder_path_p = folder_path_m + r'\peb'
    
    path_dir = folder_path_p
    if not os.path.exists(path_dir):
        print('dir of peb doesn\'t exist, a new one created!\n')
        os.makedirs(path_dir)    
    
    for file_index, file_name in enumerate(file_name_list):
        print('\ncreating plotting excel book of {}'.format(file_name))
        temperatures = temperatures_list[file_index]
        eb = Excel_Book.load_excel_book(folder_path_m, file_name)
        
        sheet_name_list = eb.wb.sheetnames
        if 'Sheet' in sheet_name_list:
            sheet_name_list.remove('Sheet')
        
        for sheet_name in sheet_name_list:
            print('\nfile_name={} \noperating {}\n'.format(file_name, sheet_name))
            es_m = Excel_Sheet(eb, sheet_name)
            for T in temperatures:
                print('{} K sheet creating.'.format(T))
                es_T = es_m.at_temperature(temperature=T, dT=0.1, folder_path=folder_path_p)
                print('{} K sheet created.'.format(T))
                es_T.save()
                print('{} K sheet saved.\n'.format(T)) 
        print('created ploting excel book of {}.\n'.format(file_name))
    print('finished!')
    

if __name__ == '__main__':
    
    
    '''
    input:  folder_path, 
            H_dir, 
            switch_channel_list, 
            para_cal_rho_cm_list, 
            mea_id_list, 
            temperatures_list,
            dT

    Returns
    -------
    None.

    '''
    
    # input folder_path
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20211125f1, LCO-LSMO0.67, RT, RHxyT\excel_copied_from_tdms\RHT\RHzT'
    folder_path_m = folder_path + r'\meb'           # measure books
    
    H_dir = 'z'

    '''
    create meb
        input folder_path
        creat measure excel books of the excel books in the folder_path. concat all the columns.
        so far measure_type='RHT,switch'  only.
        
    '''
    file_name_list= os.listdir(folder_path)
    
    path_dir = folder_path_m
    if not os.path.exists(path_dir):
        print('dir of meb doesn\'t exist, a new one created!\n')
        os.makedirs(path_dir)
        
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1] == '.xlsx':
            print('creating measure excel book of {}'.format(file_name))
            eb = Excel_Book.load_excel_book(folder_path,file_name)
            eb.write_to_measure_excel_book(d_folder_path=folder_path_m, d_file_name='', measure_type='RHT,switch')
            print('created measure excel book of {}\n'.format(file_name))
    
    
    
    '''
    insert H calculated from I(magnet).
        input folder_path_m, H_dir (I_to_H(H_dir)) H_dir = 'z' or 'xy'
        在folder_path目录下的excel文件中插入H列，由定标曲线拟合从I(magnet)转换得到。
        设置 folder_path_m, 即可插H。
    '''
    print('\ninserting H calculated from I(magnet).')
    
    # input H_dir
    # H_dir = 'z'
    
    file_name_list = os.listdir(folder_path_m)
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1]=='.xlsx':
            print('\nprocessing excel book: {}'.format(file_name))
            eb = Excel_Book.load_excel_book(folder_path_m,file_name)
            for sheet_name in eb.wb.sheetnames:
                es = Excel_Sheet(eb, sheet_name)
                if 'Lakeshore643 Output/A' in es.df_es.columns:
                    print('{}.H inserting.'.format(sheet_name))
                    es.I_to_H(H_dir)
                    es.save()
    
    
    '''
    calculate \rho & \rho_sheet of meb
    input para_cal_rho_cm_list [length, width, thickness],
    '''
    print('\ncalculating rho & rho_sheet of meb.')
    file_name_list = os.listdir(folder_path_m)
    
    switch_channel_list = [ 
                            [1,1],[1,2],[1,3],[1,4],[1,5],[1,6],[1,7],[1,8],[1,9],[1,10],
                            [2,1],[2,2],[2,3],[2,4],[2,5],[2,6],[2,7],[2,8],[2,9],[2,10],
                          ]
          
    # input para_cal_rho_cm_list: [
    #       para_cal_rho_cm_list = [
    #                                 (slot1 channel1): [length width thickness],
    #                                 ...
    #                                 (slot1 channel10): [length width thickness],
    #                                 (slot2 channel1): [length width thickness],
    #                                 ...
    #                                 (slot2 channel10): [length width thickness],
    #                             ]

    para_cal_rho_cm_list = [
                              [],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                                 
                              [],                                  

                              [3e-2, 1e-2, 2.6e-6], 
                              [],
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [], 
                              
    
                              [],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                                 
                              [],                                  

                              [3e-2, 1e-2, 2.6e-6], 
                              [],
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [3e-2, 1e-2, 2.6e-6],                                 
                              [],                               
                            ]
    
    para_list = list(zip(switch_channel_list, para_cal_rho_cm_list))  
    
    for file_name in file_name_list:
        if os.path.splitext(file_name)[1]=='.xlsx':
            print(file_name +': rho + rho_sheet')
            eb= Excel_Book.load_excel_book(folder_path_m, file_name)
            for para in para_list:
                if not para[1] == []:
                    channel_name = 'Switch-Slot%d CH%d'%(para[0][0], para[0][1])
                    
                    for sheet_name in eb.wb.sheetnames:
                        if channel_name in sheet_name:
                            es = Excel_Sheet(eb, sheet_name)
                            print(es.sheet_name)
                            l=para[1]
                            es.calculate_rho(l)
                            es.calculate_rho(l[0:2])
                            es.save()
                            
                    
                    
    '''
    create peb
        input folder_path_m, file_name_list, temperatures_list (of the temperatures in the excel books of RHT),
        
        split sheets of different T of one channel (sheet_name), and save with file_name = <'t=p, ' + self.sheet_name +'.xlsx'>, 
            sheet_name = <T={}K,{},{}'.format(temperature, self.sheet_name, self.file.file_nm[5:20])> if not specified.
            
    '''
    #input mea_id_list
    mea_id_list =   [
                        'DB20211125f1m39, RHzT',
                        'DB20211129f1m2, LCO-LSMO0.67, RHzT',
                        'DB20211129f1m5, LCO-LSMO0.67, RHzT',


                        ]
                    #     'DB20210510m27, RHT, switch, 190k-5k.xlsx',
                    #     'DB20210510m28, RHT, switch, 170k-.xlsx',
                    #     'DB20210510m33, RHT, switch, 40k-.xlsx',
                    #     'DB20210510m35, RHT, switch, .xlsx',
                        
                    # ]
    file_names = os.listdir(folder_path_m)
    file_name_list = []
    for mea_id in mea_id_list:
        for file_name in file_names:
            if mea_id in file_name:
                file_name_list.append(file_name)
    
    
    # input temperatures_list
    tl =  list(np.arange(120,301, 10))
    tl = tl.append(320)
    temperatures_list = [   [60,80,90, 100],
                            [300],
                            [310, 320],
                            ]
                        #   np.arange(320,199,-10),
                        #   [190,180],
                        #   np.arange(170,39,-10),
                        #   np.arange(40,9,-10),
                        #   [320],                            
                        # ]
    
    folder_path_p = folder_path_m + r'\peb'
    
    path_dir = folder_path_p
    if not os.path.exists(path_dir):
        print('dir of peb doesn\'t exist, a new one created!\n')
        os.makedirs(path_dir)    
    
    for file_index, file_name in enumerate(file_name_list):
        print('\ncreating plotting excel book of {}'.format(file_name))
        temperatures = temperatures_list[file_index]
        eb = Excel_Book.load_excel_book(folder_path_m, file_name)
        
        sheet_name_list = eb.wb.sheetnames
        if 'Sheet' in sheet_name_list:
            sheet_name_list.remove('Sheet')
        
        for sheet_name in sheet_name_list:
            print('\nfile_name={} \noperating {}\n'.format(file_name, sheet_name))
            es_m = Excel_Sheet(eb, sheet_name)
            for T in temperatures:
                print('{} K sheet creating.'.format(T))
                es_T = es_m.at_temperature(temperature=T, dT=0.1, folder_path=folder_path_p)
                print('{} K sheet created.'.format(T))
                es_T.save()
                print('{} K sheet saved.\n'.format(T)) 
        print('created ploting excel book of {}.\n'.format(file_name))
    print('finished!')
    
    
    