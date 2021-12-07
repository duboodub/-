# -*- coding: utf-8 -*-
"""
Created on Sat Dec  4 03:59:00 2021

@author: db

input RHT excel book 

every es of T
    catch T
    creat es1, [R(Imin), R(Imax), mean(Imin), mean(Imax), std(Imin), std(), RH, RH_err,  ]
    


"""

import matplotlib.pyplot as plt 
import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np
import os, re
from data_process_oop import Excel_Book, Excel_Sheet, File


def tryint(s):                       
    try:
        return int(s)
    except ValueError:
        return s

def str2int(v_str):                
    return [tryint(sub_str) for sub_str in re.split('([0-9]+)', v_str)]


def sort_humanly(v_list):   
    # sort humanly
    return sorted(v_list, key=str2int)



if __name__ == '__main__':
    
    '''
        磁场方向，电压方向
    '''
    folder_path = r'E:\NAS, data\Data Temporary\Cryostat B\DUBO\DB20210601f1, LSMO.26nm, RHzT\tdms\excel_copied_from_tdms'
    file_name = 'bt=peb, Switch-Slot1 CH3,t=m,.xlsx'
    file_name_RH_T = 'RH_T, '+ file_name
    
    sheet_name_RH_T = 'RH_T'
    eb = Excel_Book.load_excel_book(folder_path,file_name)
    
    eb1= Excel_Book(folder_path, file_name_RH_T)
    eb1.save()
    
    df2 = pd.DataFrame()
    
    sheetnames = sort_humanly(eb.wb.sheetnames)
    
    for sheet_name in sheetnames:
        if 'T=' in sheet_name:
            print(sheet_name+' ing')
            es = Excel_Sheet( eb, sheet_name = sheet_name)
            
            col_name_R = 'Resistance(Ω) '
            col_name_I = 'Lakeshore643 Output/A'
            
            df1 = pd.DataFrame(columns=['R(Imin)', 'R(Imax)', 'mean(Imin)', 'mean(Imax)', 'H(Imin)', 'H(Imax)', 'std(Imin)', 'std(Imax)', 'RH', 'RH_err', ])
            df1 = pd.DataFrame()
            col_names = { 'T':'T', 'R(%s)':'R(%s)', 'H(%s)':'H(%s)', 'mean(%s)':'mean(%s)', 'std(%s)':'std(%s)', 'RH':'RH', 'mean_RH':'mean_RH', 'std_RH':'std_RH', }
            Is = {'Imin': -60, 'Imax': 60}
            for I in Is:
                
                df= es.df_es[ es.df_es[col_name_I]==Is[I] ]
                H = df['H(from I)/Gauss'].iloc[1]


                s= pd.Series(data= df[col_name_R], name= col_names['R(%s)']%I,)
                s= s.reset_index(drop=True)
                df1= pd.concat([df1, s], axis=1, ignore_index=False,)
                
                s= pd.Series(data=np.array(H), name= col_names['H(%s)']%I,)
                s= s.reset_index(drop=True)
                df1= pd.concat([df1, s], axis=1, ignore_index=False,)
 
                s= pd.Series(data=np.array( df1[col_names['R(%s)']%I].mean() ), name= col_names['mean(%s)']%I,)
                s= s.reset_index(drop=True)
                df1= pd.concat([df1, s], axis=1, ignore_index=False,)
 
                s= pd.Series(data=np.array( df1[col_names['R(%s)']%I].std() ), name= col_names['std(%s)']%I,)
                s= s.reset_index(drop=True)
                df1= pd.concat([df1, s], axis=1, ignore_index=False,)
                
                # ac = pd.Series(np.array(H) )
                # df1['H(%s)'%I] = ac
                # df1['R(%s)'%I] = df[col_name_R]
                # ab =pd.Series( [df1['R(%s)'%I].mean()] )
                # df1['mean(%s)'%I] = ab
                # bc= pd.Series( [df1['R(%s)'%I].std()] )
                # df1['std(%s)'%I] = bc
            
            
            RH=[]
            dH = (df1['H(Imax)'][0]-df1['H(Imin)'][0])
            for R0 in df1[col_names['R(%s)']%list(Is)[0]].dropna():
                for R1 in df1[col_names['R(%s)']%list(Is)[1]].dropna():
                    # rh = (R1-R0) / (Is.values(1)-Is.values(0))
                    rh = (R1-R0) / dH
                    RH.append(rh,)
            
            RH = pd.Series(data= RH, name= col_names['RH'],)
            RH= RH.reset_index(drop=True)
            df1= pd.concat([df1, RH], axis=1, ignore_index=False,)      
            
            
            s= pd.Series(data=df1[col_names['RH']].mean() , name= col_names['mean_RH'],)
            s= s.reset_index(drop=True)
            df1= pd.concat([df1, s], axis=1, ignore_index=False,)
 
            s= pd.Series(data=df1[col_names['RH']].std() , name= col_names['std_RH'],)
            s= s.reset_index(drop=True)
            df1= pd.concat([df1, s], axis=1, ignore_index=False,)   
            
            T = [int(re.findall(r'\d+', sheet_name)[0])]
            s= pd.Series(data=T , name= col_names['T'],)
            s= s.reset_index(drop=True)
            df1.insert(0, s.name, s)
            
            es1 = Excel_Sheet(data=df1, sheet_name=sheet_name, file= File(folder_path, file_name_RH_T))
            es1.save()
            
            df2= pd.concat([ df2, df1[[ col_names['T'], col_names['mean_RH'], col_names['std_RH'] ]].dropna() ])
    
    
    es2 = Excel_Sheet(data=df2, sheet_name=sheet_name_RH_T, file= File(folder_path, file_name_RH_T) )
    es2.save()

            
    
    
    
    
    
    
    
            
                      
            