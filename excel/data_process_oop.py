# -*- coding: utf-8 -*-
"""
Created on Mon Jun  7 16:28:01 2021

@author: db
"""

import matplotlib.pyplot as plt 
import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np
import os

class File(object):
    
    def __init__(self, folder_path='', file_name=''):
        self.folder_path = folder_path
        self.file_name = file_name
        # @property
        self.file_nm = os.path.splitext(self.file_name)[0]
        self.file_post = os.path.splitext(self.file_name)[1]
        self.file_path = os.path.join(self.folder_path, self.file_name)
        
        
    def change(self,folder_path='', file_name=''):
        if folder_path == '':
            folder_path = self.folder_path
        if file_name == '':
            file_name = self.file_name
        
        file = File(folder_path, file_name)
        return file
        
    
    
    def change_directory(self):
        pass
    
    def change_file_name(self):
        pass
    
    
class Excel_Book(object):
    
    def __init__(self, folder_path, file_name):
        # super().__init__()
        self.file = File(folder_path, file_name)
        self.wb = Workbook()
        
        
    # 作subclass 可用
    # def load_Excel_Book():
        # eb = Excel_Book()
        
   
    def load_excel_book(folder_path, file_name, read_only:bool = False):
        eb = Excel_Book(folder_path,file_name,)
        eb.wb = load_workbook(eb.file.file_path, read_only)
        
        return eb
        
    ##
    def create_excel_book(self, folder_path, file_name, r_only:bool = False):
        p = os.path.join(folder_path, file_name)
        if os.path.exists(folder_path):
            print('exist')
            
        else:
            os.makedirs(folder_path)
            wb = Workbook()
            wb.save(p) 
        eb = Excel_Book(folder_path,file_name,)
        return eb
    
    
    def save(self, folder_path= '', file_name='', ):
        if not folder_path == '':
            self.file.folder_path = folder_path
        if not file_name == '':
            self.file.file_name = file_name
        ##
        if os.path.exists(folder_path):
            print('folder_path %s exist\n'%folder_path )
            
        self.wb.save(self.file.file_path)
        #to check the path
        #to save to other path
        
    
    def pull_excel_sheet(self, sheet_name,): 
        pd.read_excel(io=self.file.file_name, sheet_name=sheet_name,)
        pass 
    
    # def write_to_measure_excel_book(self, )
            
      
    #switch
    def write_to_measure_excel_book(self, d_folder_path='', d_file_name='', measure_type:str=''):
        '''
        
        
        Parameters
        ----------
        d_folder_path : TYPE, optional
            DESCRIPTION. The default is '', d_folder_path = self.file.folder_path.
        d_file_name : TYPE, optional
            DESCRIPTION. The default is '', d_file_name = 't=m, '+self.file.file_name.
        measure_type : str, optional
            DESCRIPTION. The default is '', measure_type = 'RHT,switch'.

        Raises
        ------
        ValueError
            DESCRIPTION.

        Returns
        -------
        None.

        '''
        s_folder_path = self.file.folder_path
        s_file_name = self.file.file_name
        
        if d_folder_path=='':
            d_folder_path = self.file.folder_path
        if d_file_name == '':
            d_file_name = 't=m, '+self.file.file_name
        d_file = File(d_folder_path,d_file_name)
        
        if measure_type == 'RHT,switch':
            sheet_name_source_list = ['system time', 'magnet data', ]
            sheet_name_new_list = []
            
            eb = Excel_Book.load_excel_book(s_folder_path, s_file_name)
            for sheet_name in eb.wb.sheetnames:
                if 'Switch-Slot' in sheet_name:
                    sheet_name_new_list.append(sheet_name)  
            
            
            for sheet_name_new in sheet_name_new_list:
                print('measure excel sheet creating: {}.\r'.format(sheet_name_new))
                es2 = Excel_Sheet(eb, sheet_name = sheet_name_new, file=d_file)

                for sheet_name_source in sheet_name_source_list:
                    es1 = Excel_Sheet( eb, sheet_name = sheet_name_source)
                    for col_name in es1.df_es.columns:
                        series_s = es1.df_es[col_name]
                        es2.df_es.insert(0,series_s.name, series_s)
                        
                print('measure excel sheet created.\r')
                es2.save(es2.file, sheet_name='%s,t=m,'%sheet_name_new)
                        
            
        else:
            raise ValueError('measure_type defined wrongly!')
                            
       
class Excel_Sheet(object):
    
    def __init__(self, data, sheet_name:str, file:File=None,  ):
        
        self.df_es = None
        self.file:File = None
        self.sheet_name = sheet_name
            
        # eb, sheet_name,
        if isinstance(data, Excel_Book) :
            if file==None:
                self.file = data.file
            else:
                self.file = file
                
            self.df_es = pd.read_excel(data.file.file_path, sheet_name=sheet_name,)
            self.sheet_name = sheet_name
        
        # df, 
        elif isinstance(data,pd.DataFrame):
            if not (file==None):
                self.file = file
                self.df_es = data
                self.sheet_name = sheet_name
                 
            else:
                raise ValueError(' "file" should be defined, when "data" is DataFrame type.\n')
        
        # df, sheet_name, file_path
        
         
        
    def save(self, file:File=None, sheet_name=''):
        '''
        if "file" or "sheet_name" defined, change self.file & self.sh
        将df保存到file_path[sheet_name]中，if file_path not exist, creat one, if sheet_name exists, overwrite it.
        
        Parameters
        ----------
        file : File, optional
            DESCRIPTION. The default is None.
        sheet_name : TYPE, optional
            DESCRIPTION. The default is ''.

        Returns
        -------
        None.

        '''
        # 是否改地址
        if not file == None:
            self.file=file
                    
            
        if sheet_name == '':
            sheet_name = self.sheet_name
        
        df = self.df_es
        
        
        if not os.path.exists(self.file.file_path):
            wb = Workbook()
            wb.save(self.file.file_path)
            print('excel_path doesn\'t exist! So a new file created.\r')
            
        wb = load_workbook(self.file.file_path)    
        
        if sheet_name not in wb.sheetnames:
            with pd.ExcelWriter(self.file.file_path,engine= 'openpyxl') as writer:
                writer.book = wb
                # writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
                
                # wb = load_workbook('sample_file.xlsx')
                # with pd.ExcelWriter('sample_file.xlsx',engine= 'openpyxl') as writer:
                # 如两个workbook一样则会新创一个sheet，如不一样这直接修改
                    
                df.to_excel(writer, sheet_name, index= False)
                #如sheet name冲突，则会新创一个sheet
                
                writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)    
                writer.save()
                
        # protect the data
        # else:
        #     print('{}[{}] exists, so passed.\n'.format(excel_path,sheet_name))
        
        else:
            with pd.ExcelWriter(self.file.file_path,engine= 'openpyxl') as writer:
                writer.book = wb
                wb.remove_sheet(wb[sheet_name])
                df.to_excel(writer, sheet_name, index= False)
                writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)  
                writer.save()
        
        if 'Sheet' in wb.sheetnames:
            with pd.ExcelWriter(self.file.file_path,engine= 'openpyxl') as writer:
                writer.book = wb
                wb.remove_sheet(wb['Sheet'])
                writer.save()
                
        print('saved. ')
            
    
    def at_temperature(self, temperature, dT=1, folder_path='', file_name='', sheet_name='',):
        '''
        return a excel_sheet of temperature +- dT

        Parameters
        ----------
        temperature : TYPE
            DESCRIPTION.
        dT : TYPE, optional
            DESCRIPTION. The default is 1.
        folder_path : TYPE, optional
            DESCRIPTION. The default is ''.
        file_name : TYPE, optional
            DESCRIPTION. The default is ''.
        sheet_name : TYPE, optional
            DESCRIPTION. The default is ''.

        Returns
        -------
        es1 : TYPE
            DESCRIPTION.

        '''
        col_name_T = 'sample temperature/K'
        
        if folder_path == '':
            folder_path= self.file.folder_path
        
        if file_name == '':
            file_name = 't=p, ' + self.sheet_name +'.xlsx'
            
        if sheet_name == '':
            sheet_name = 'T={}K,{}'.format(temperature, self.sheet_name, ) #self.file.file_nm[11:20]
        
        file= File(folder_path,file_name)
            
        df = self.df_es[(self.df_es[col_name_T] > (temperature-dT)) & (self.df_es[col_name_T] < (temperature+dT))]
        
        es1 = Excel_Sheet(df, file=file, sheet_name = sheet_name, )
        
        return es1
        
        
    def calculate_rho(self, l_w_t_list ):
        '''
        calculate_rho
        
        'Resistance' in col_name_R
        
        Parameters
        ----------
        l_w_t_list : TYPE
            l_w_t = length_width_(thickness)

        Raises
        ------
        ValueError
            DESCRIPTION.

        Returns
        -------
        rho : TYPE
            DESCRIPTION.

        '''
        if len(l_w_t_list) == 3:
            length = l_w_t_list[0]
            width = l_w_t_list[1]
            thickness = l_w_t_list[2]
            
        elif len(l_w_t_list) == 2:
            length = l_w_t_list[0]
            width = l_w_t_list[1]
            thickness = None
        else:
            raise ValueError('please input a correct l_w_t_list.')
            
        col_name_R = ''
        for col_name in self.df_es.columns:
            if 'Resistance' in col_name:
                col_name_R = col_name
                    
        allow_duplicates = True
        if thickness == None:
            if isinstance(length,(int,float)) and isinstance(width,(int,float)):
                rho = self.df_es[col_name_R]*width/length
                self.df_es.insert(len(self.df_es.columns),'rho_sheet/Ohm', rho, allow_duplicates)
                return rho
                
            else:
                raise ValueError('length or/and width should be int/float.')
        
        elif isinstance(thickness,(int,float)):
            if isinstance(length,(int,float)) and isinstance(width,(int,float)):
                rho = self.df_es[col_name_R]*width*thickness/length
                self.df_es.insert(len(self.df_es.columns),'rho/(Ohm)cm', rho, allow_duplicates)
                return rho
                
            else:
                raise ValueError('length or/and width should be int/float.')  
        
        else:
            raise ValueError('thickness should be int/float.') 
        
            
    def I_to_H(self, H_dir:str):
        '''
        insert a column of H/Gauss calculated from I(magnet) in the es.

        Parameters
        ----------
        H_dir : str
            H_dir == 'z':
                the H is applied in the direction of 'z'
                
            H_dir == 'xy':
                the H is applied in the direction of 'xy'
                
                
        Raises
        ------
        ValueError
            DESCRIPTION.

        Returns
        -------
        None.

        '''
        self.df_es.columns
        col_name = 'Lakeshore643 Output/A'
        I_col = self.df_es[col_name]
        if H_dir == 'xy':
            H_col = 89.7816 * I_col
            
        elif H_dir == 'z':
            # original
            # Intercept = 2.79202
            # B1 = -94.16396
            # B2 = -0.02314
            # B3 = -0.00106
            # B4 = -3.70453E-6
            # B5 = 7.28664E-7
            # B6 = 2.8925E-9
            # B7 = 3.05798E-10
            # B8 = -3.77237E-13
            # B9 = -5.98534E-14
            # H_col = Intercept + B1*I_col**1 + B2*I_col**2 + B3*I_col**3 + B4*I_col**4 + B5*I_col**5 + B6*I_col**6 + B7*I_col**7 + B8*I_col**8 + B9*I_col**9
            
            Intercept = 2.79202
            B1 = 94.16396
            B2 = -0.02314
            B3 = 0.00106
            B4 = -3.70453E-6
            B5 = -7.28664E-7
            B6 = 2.8925E-9
            B7 = -3.05798E-10
            B8 = -3.77237E-13
            B9 = 5.98534E-14
            H_col = Intercept + B1*I_col**1 + B2*I_col**2 + B3*I_col**3 + B4*I_col**4 + B5*I_col**5 + B6*I_col**6 + B7*I_col**7 + B8*I_col**8 + B9*I_col**9
            
        else:
            raise ValueError('The H direction should be indicated rightly.')
            
        self.df_es.insert(0,'H(from I)/Gauss', H_col)
        
        
            

class measure_excel(File):
    
    def __init__(self, folder_path, file_name):
        File.__init__(self, folder_path, file_name)
        self.measure_id = self.find_measure_id()
        
        
        if self.file_post != 'xlsx':
            print('It\'s not a excel, you can\'t assign this file as measure_excel!')
            return 1
         
        
        
    def find_measure_id(self):
        s0 = self.file_name.find('DB')
        s1 = s0 + 8
        measure_id = self.file_name[s0:s1]
        return measure_id
    
class data_sheet(measure_excel):
    def __init__(self, folder_path, file_name, sheet_name):
        measure_excel.__init__(self, folder_path, file_name)
        self.sheet_name = sheet_name
        self.data = self.to_df()
        
    def to_df(self):
        pd.read_excel(self.excel_path, sheet_name=self.sheet_name, engine='openpyxl',)
        

class ExcelBook(Workbook):
    
    def __init__(self, wb_path, ):
        pass
          

if __name__ == '__main__':
    
    es = Excel_Sheet(1,34,55)
    es.save()
                    
        
            
    eb = Excel_Book('asdfds','dfjie', )
    # print(eb.file.file_name)
